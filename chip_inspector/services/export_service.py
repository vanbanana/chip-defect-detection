"""
Export service - Data export functionality.
Supports exporting inspection results to various formats.
"""
import csv
import json
from typing import List, Optional
from pathlib import Path
from datetime import datetime

from core.models import InspectionResult
from core.enums import ExportFormat
from core.exceptions import ExportError
from utils.logger import get_logger


class ExportService:
    """Service for exporting inspection results."""

    def __init__(self):
        self._logger = get_logger(__name__)

    def export_to_csv(
        self,
        results: List[InspectionResult],
        output_path: str,
        include_headers: bool = True
    ) -> int:
        """
        Export inspection results to CSV format.

        Args:
            results: List of inspection results
            output_path: Path to output CSV file
            include_headers: Whether to include column headers

        Returns:
            Number of rows written
        """
        try:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)

                # Headers
                if include_headers:
                    headers = [
                        'ID',
                        'Timestamp',
                        'Image Name',
                        'Status',
                        'Defect Area',
                        'Recipe',
                        'Processing Time (ms)',
                        'Image Path'
                    ]
                    writer.writerow(headers)

                # Data rows
                for result in results:
                    row = [
                        result.id or '',
                        result.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                        result.image_name,
                        result.get_status_display(),
                        f"{result.defect_area:.2f}",
                        result.recipe_name,
                        f"{result.processing_time_ms:.2f}",
                        result.image_path
                    ]
                    writer.writerow(row)

            count = len(results)
            self._logger.info(f"Exported {count} results to CSV: {output_path}")
            return count

        except Exception as e:
            self._logger.error(f"CSV export failed: {str(e)}")
            raise ExportError(f"Failed to export CSV: {str(e)}")

    def export_to_json(
        self,
        results: List[InspectionResult],
        output_path: str,
        pretty: bool = True
    ) -> int:
        """
        Export inspection results to JSON format.

        Args:
            results: List of inspection results
            output_path: Path to output JSON file
            pretty: Whether to format JSON with indentation

        Returns:
            Number of results exported
        """
        try:
            data = {
                'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_count': len(results),
                'results': [r.to_dict() for r in results]
            }

            with open(output_path, 'w', encoding='utf-8') as f:
                if pretty:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                else:
                    json.dump(data, f, ensure_ascii=False)

            self._logger.info(f"Exported {len(results)} results to JSON: {output_path}")
            return len(results)

        except Exception as e:
            self._logger.error(f"JSON export failed: {str(e)}")
            raise ExportError(f"Failed to export JSON: {str(e)}")

    def export_to_excel(
        self,
        results: List[InspectionResult],
        output_path: str
    ) -> int:
        """
        Export inspection results to Excel format.

        Args:
            results: List of inspection results
            output_path: Path to output Excel file

        Returns:
            Number of results exported
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inspection Results"

            # Headers
            headers = [
                'ID',
                'Timestamp',
                'Image Name',
                'Status',
                'Defect Area',
                'Recipe',
                'Processing Time (ms)',
                'Image Path'
            ]

            # Header style
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data rows
            for row, result in enumerate(results, 2):
                ws.cell(row=row, column=1).value = result.id
                ws.cell(row=row, column=2).value = result.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=3).value = result.image_name
                ws.cell(row=row, column=4).value = result.get_status_display()
                ws.cell(row=row, column=5).value = result.defect_area
                ws.cell(row=row, column=6).value = result.recipe_name
                ws.cell(row=row, column=7).value = result.processing_time_ms
                ws.cell(row=row, column=8).value = result.image_path

                # Color code status
                status_cell = ws.cell(row=row, column=4)
                if result.is_ok:
                    status_fill = PatternFill(start_color='00C853', end_color='00C853', fill_type='solid')
                else:
                    status_fill = PatternFill(start_color='D50000', end_color='D50000', fill_type='solid')
                status_cell.fill = status_fill
                status_cell.font = Font(color='FFFFFF', bold=True)

            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Save
            wb.save(output_path)

            self._logger.info(f"Exported {len(results)} results to Excel: {output_path}")
            return len(results)

        except ImportError:
            self._logger.error("openpyxl not installed")
            raise ExportError("Excel export requires openpyxl package. Install with: pip install openpyxl")
        except Exception as e:
            self._logger.error(f"Excel export failed: {str(e)}")
            raise ExportError(f"Failed to export Excel: {str(e)}")

    def export(
        self,
        results: List[InspectionResult],
        output_path: str,
        format: Optional[ExportFormat] = None
    ) -> int:
        """
        Export inspection results with automatic format detection.

        Args:
            results: List of inspection results
            output_path: Path to output file
            format: Export format, or None to auto-detect from extension

        Returns:
            Number of results exported
        """
        if format is None:
            ext = Path(output_path).suffix.lower()
            format_map = {
                '.csv': ExportFormat.CSV,
                '.json': ExportFormat.JSON,
                '.xlsx': ExportFormat.EXCEL
            }
            format = format_map.get(ext, ExportFormat.CSV)

        if format == ExportFormat.CSV:
            return self.export_to_csv(results, output_path)
        elif format == ExportFormat.JSON:
            return self.export_to_json(results, output_path)
        elif format == ExportFormat.EXCEL:
            return self.export_to_excel(results, output_path)
        else:
            raise ExportError(f"Unsupported export format: {format}")

    def export_statistics_report(
        self,
        results: List[InspectionResult],
        output_path: str
    ) -> None:
        """
        Export a statistics report.

        Args:
            results: List of inspection results
            output_path: Path to output text file
        """
        try:
            from services.result_service import ResultService
            result_service = ResultService()
            stats = result_service.get_batch_statistics(results)

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 50 + "\n")
                f.write("检测统计报告\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

                f.write("总体统计:\n")
                f.write("-" * 30 + "\n")
                f.write(f"  总数:     {stats['total']}\n")
                f.write(f"  合格:     {stats['ok_count']}\n")
                f.write(f"  不良:     {stats['ng_count']}\n")
                f.write(f"  良率:     {stats['pass_rate']:.2f}%\n\n")

                f.write("缺陷统计:\n")
                f.write("-" * 30 + "\n")
                f.write(f"  平均面积: {stats['avg_area']:.2f} px²\n")
                f.write(f"  最大面积: {stats['max_area']:.2f} px²\n\n")

                f.write("性能统计:\n")
                f.write("-" * 30 + "\n")
                f.write(f"  平均耗时: {stats['avg_time_ms']:.2f} ms\n")

            self._logger.info(f"Statistics report exported: {output_path}")

        except Exception as e:
            self._logger.error(f"Failed to export statistics: {str(e)}")
            raise ExportError(f"Failed to export statistics: {str(e)}")
